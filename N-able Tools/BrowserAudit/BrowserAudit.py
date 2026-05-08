#!/usr/bin/env python3
"""
Browser CVE Audit Report Generator v6
====================================
1. Click "Select Task Report CSV(s)" and pick your CSV file(s)
2. Click "Select Device Inventory" and pick the XLSX
3. Click "Generate Report"

Install dependencies once:
    pip install pandas openpyxl xlrd
"""

import re
import threading
import tkinter as tk
from datetime import datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ── Excel helpers ─────────────────────────────────────────────────────────────
COL = {
    # Header / accent colours (used on header rows and summary tiles)
    "dark": "2C3E50", "navy": "1A5276", "red": "C0392B",
    "amber": "E67E22", "green": "1E8449", "gold": "7D6608",
    # Data-row highlight colours — all light pastels for readable black text
    "row_dual32":   "FFD0D0",   # light red   — dual 32+64-bit install
    "row_only32":   "FFF0CC",   # light amber — 32-bit only install
    "row_peruser":  "D0E8FF",   # light blue  — per-user/AppData install
    "row_risk":     "FDEBD0",   # light orange — any risk row on browser sheets
    "row_blue":     "EBF5FB",   # light blue  — alternating clean row
    "row_gold":     "FEF9E7",   # light gold  — alternating row (stale sheet)
    "row_gold2":    "F9F3D2",   # light gold2 — alternating row (stale sheet)
    "white":        "FDFEFE",   # near-white  — alternating clean row
}

def _fill(h):
    return PatternFill("solid", start_color=h)

HDR_FONT  = Font(name="Arial", bold=True, color="FFFFFF", size=10)
BODY_FONT = Font(name="Arial", size=9)
WRAP      = Alignment(wrap_text=True, vertical="top")
CENTER    = Alignment(horizontal="center", vertical="center")
_thin     = Side(style="thin", color="BDBDBD")
BORDER    = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)

def _hdr(ws, row, ncols, fill):
    for c in range(1, ncols + 1):
        cell = ws.cell(row, c)
        cell.fill = fill; cell.font = HDR_FONT; cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[row].height = 22

def _widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

def _write_legend(ws, start_row, entries):
    """
    Write a small colour legend below the data.
    entries: list of (hex_colour, label_text)
    Writes two rows: a blank spacer then one row per entry.
    """
    LBL_FONT = Font(name="Arial", size=8, bold=True, color="404040")
    spacer = start_row
    ws.row_dimensions[spacer].height = 8
    for i, (hex_col, label) in enumerate(entries, spacer + 1):
        swatch = ws.cell(i, 1)
        swatch.fill = _fill(hex_col)
        swatch.border = BORDER
        swatch.value = ""
        ws.row_dimensions[i].height = 16
        lbl = ws.cell(i, 2, f"  {label}")
        lbl.font = LBL_FONT
        lbl.alignment = Alignment(vertical="center")


# ── Browser parsing ───────────────────────────────────────────────────────────
# N-able output can wrap/truncate the InstallScope column, for example:
#   P..., Per, Per-User, User, S..., Sys, System
# It can also contain paths with spaces such as "Program Files (x86)".
# So parse from Browser -> path ending in .exe -> version -> architecture -> optional scope.
_BPAT = re.compile(
    r"(Google Chrome|Microsoft Edge|Mozilla Firefox|Brave|Opera)"
    r"\s+(C:\\.*?\.exe)"
    r"\s+([\d.]+)"
    r"\s+(32-bit|64-bit)"
    r"(?:\s+([A-Za-z.\-]+))?",
    re.I,
)

def _normalise_scope(scope, path):
    s = (scope or "").strip().lower()
    p = (path or "").lower()

    # AppData path is the strongest signal for per-user install, even when
    # N-able truncates InstallScope to P... or omits the full value.
    if "\\appdata\\" in p or "appdata" in p:
        return "Per-User"
    if s.startswith(("p", "per", "user")):
        return "Per-User"
    if s.startswith(("s", "sys")):
        return "System"
    return "Unknown"

def _resolve_arch(reported_arch, path, browser):
    """
    Correct N-able architecture mis-reporting using install path as ground truth.

    Rule 1 — Chrome in Program Files (x86) -> 32-bit regardless of N-able's report.
      Chrome is deployed via Intune as a 64-bit MSI which always installs to
      Program Files (not x86). If it lands in (x86) the Intune deployment failed
      and the device has a legacy or manually installed 32-bit copy. Treat it as
      the issue it is rather than trusting N-able's PE-header read.

    Rule 2 — Any browser in Program Files (no (x86)) reported as 32-bit -> 64-bit.
      A 32-bit binary cannot physically live in Program Files; Windows redirects
      32-bit installers to Program Files (x86) at install time. N-able is wrong.

    Anything else -> trust N-able.
    """
    p    = path.lower()
    x86  = "(x86)" in p

    if browser == "Google Chrome" and x86:
        return "32-bit"   # Chrome in (x86) = Intune deployment issue, treat as 32-bit

    if reported_arch == "32-bit":
        in_pf = "\\program files\\" in p or p.startswith("c:\\program files\\")
        if in_pf and not x86:
            return "64-bit"   # path proves it cannot be 32-bit; N-able is wrong

    return reported_arch


def parse_browsers(output):
    results = []
    text = str(output).replace("\r", "\n").replace(";", "\n")
    for m in _BPAT.finditer(text):
        path  = m.group(2).strip()
        scope = _normalise_scope(m.group(5), path)
        arch  = _resolve_arch(m.group(4).strip(), path, m.group(1).strip())
        results.append({
            "browser": m.group(1).strip(),
            "path":    path,
            "version": m.group(3).strip(),
            "arch":    arch,
            "scope":   scope,
        })
    return results


# ── Firefox last-used parsing ─────────────────────────────────────────────────
# Matches lines in the "Firefox Usage Report" section produced by Get-FirefoxLastUsed.
# Expected line format (tab or multi-space separated):
#   <UserName>   <ProfileDirName>   <YYYY-MM-DD HH:MM:SS>
# The datetime is captured loosely so minor PS formatting differences don't break it.
_FF_USED_PAT = re.compile(
    r"^[ \t]*(\S+)"                       # Windows username (no spaces)
    r"[ \t]+"
    r"(\S+)"                              # Firefox profile dir name
    r"[ \t]+"
    r"(\d{1,2}/\d{1,2}/\d{4}"            # date variants:  M/D/YYYY
    r"|\d{4}-\d{2}-\d{2}"                #                 YYYY-MM-DD
    r"|\d{1,2}\s+\w+\s+\d{4})"          #                 D Month YYYY
    r"[ \t]+(\d{1,2}:\d{2}(?::\d{2})?(?:\s*[AP]M)?)",  # time, optional seconds/AM-PM
    re.MULTILINE | re.IGNORECASE,
)

def parse_firefox_last_used(output):
    """
    Extract Firefox last-used records from a PS script output string.
    Returns a list of dicts: {user, profile, last_used_raw, last_used_dt}.
    Only the block after 'Firefox Usage Report:' is searched so that
    the browser-install table above it doesn't produce false matches.
    """
    text = str(output).replace("\r\n", "\n").replace("\r", "\n").replace(";", "\n")

    # Narrow to the Firefox Usage Report section if the marker is present
    marker_idx = text.lower().find("firefox usage report")
    if marker_idx != -1:
        text = text[marker_idx:]

    results = []
    for m in _FF_USED_PAT.finditer(text):
        raw_dt = f"{m.group(3)} {m.group(4)}".strip()
        dt = None
        for fmt in ("%d/%m/%Y %I:%M:%S %p", "%d/%m/%Y %I:%M %p",
                    "%d/%m/%Y %H:%M:%S",    "%d/%m/%Y %H:%M",
                    "%m/%d/%Y %I:%M:%S %p", "%m/%d/%Y %I:%M %p",
                    "%m/%d/%Y %H:%M:%S",    "%m/%d/%Y %H:%M",
                    "%Y-%m-%d %H:%M:%S",    "%Y-%m-%d %H:%M",
                    "%d %B %Y %H:%M:%S",    "%d %B %Y %H:%M"):
            try:
                dt = datetime.strptime(raw_dt, fmt)
                break
            except ValueError:
                continue
        results.append({
            "user":         m.group(1),
            "profile":      m.group(2),
            "last_used_raw": raw_dt,
            "last_used_dt":  dt,
        })
    return results

def _path_note(arch, path):
    """Return a plain-English note when a 64-bit binary lives in a (x86) folder.
    Edge always installs to Program Files (x86) regardless of bitness — without
    this note, techs see (x86) in the path and incorrectly assume it is 32-bit."""
    if arch == "64-bit" and "(x86)" in path:
        return "64-bit binary in (x86) folder — expected install location for this browser"
    return ""


def detect_issues(browsers):
    """
    Returns (dual_32, only_32, per_user) — three lists of installs, each an issue on a Win11 64-bit fleet.

    dual_32  : 32-bit installs where the same browser ALSO has a 64-bit install present.
               Remediation: remove the 32-bit copy.
    only_32  : 32-bit installs where NO 64-bit copy of that browser exists.
               Remediation: replace with the 64-bit installer.
    per_user : installs under AppData / marked Per-User scope.
               Remediation: reinstall to Program Files as System scope.
    """
    by_name = {}
    for b in browsers:
        by_name.setdefault(b["browser"], []).append(b)

    dual_32, only_32 = [], []
    for name, installs in by_name.items():
        archs = {b["arch"] for b in installs}
        for b in installs:
            if b["arch"] != "32-bit":
                continue
            if "64-bit" in archs:
                dual_32.append(b)   # 32-bit alongside a 64-bit copy
            else:
                only_32.append(b)   # no 64-bit version present at all

    per_user = [b for b in browsers if "AppData" in b["path"] or b["scope"] == "Per-User"]
    return dual_32, only_32, per_user


# ── Core logic ────────────────────────────────────────────────────────────────
def load_inventory(path, stale_days):
    inv = pd.read_excel(path)
    inv.columns = inv.columns.str.strip()
    inv["Device name"] = inv["Device name"].astype(str).str.strip()
    inv["last_response_dt"] = pd.to_datetime(
        inv["Last response (Local time)"], format="%m/%d/%y %I:%M:%S %p", errors="coerce")
    today = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    inv["days_since"] = (today - inv["last_response_dt"]).dt.days
    stale = inv["days_since"] > stale_days
    client_name = inv["Customer name"].dropna().iloc[0] if "Customer name" in inv.columns and not inv["Customer name"].dropna().empty else ""
    return inv[~stale].copy(), inv[stale].copy(), str(client_name).strip()

def load_audit(paths):
    combined = pd.concat([pd.read_csv(p) for p in paths], ignore_index=True)
    combined.columns = combined.columns.str.strip()

    required = {"Task", "Status", "Device", "Date", "Output", "Client", "Site"}
    missing = required - set(combined.columns)
    if missing:
        raise ValueError(f"Task report CSV is missing required column(s): {', '.join(sorted(missing))}")

    combined["Device"] = combined["Device"].astype(str).str.strip()
    combined["Task"] = combined["Task"].astype(str).str.strip()
    combined["Status"] = combined["Status"].astype(str).str.strip()

    audit = combined[
        combined["Task"].str.contains("BrowserAudit|SimpleBrowserAudit", case=False, na=False) &
        combined["Status"].str.casefold().eq("stopped")
    ].copy()

    if audit.empty:
        raise ValueError("No stopped BrowserAudit/SimpleBrowserAudit task rows were found in the selected CSV.")

    # N-able exports dates as text such as '6 May 2026 14:51'. Parse where possible so the latest
    # result per device is selected correctly instead of relying on string sorting.
    audit["_date_dt"] = pd.to_datetime(audit["Date"], errors="coerce", dayfirst=True)
    audit = audit.sort_values(["_date_dt", "Date"], na_position="first").drop_duplicates(subset="Device", keep="last")
    audit = audit.drop(columns=["_date_dt"])
    return audit

def build_flagged(audit, inv_active):
    """One row per flagged install — dual_32 / only_32 / per_user, all filterable."""
    flagged = []
    idx = inv_active.set_index("Device name")
    seen = set()  # (device, browser, path, issue_type) — prevent double-counting
    for _, row in audit.iterrows():
        device = row["Device"]
        ir = idx.loc[device] if device in idx.index else None
        base = {
            "Device":        device,
            "Client":        ir["Customer name"]               if ir is not None and "Customer name"             in ir.index else row.get("Client", ""),
            "Site":          ir["Site name"]                   if ir is not None and "Site name"                 in ir.index else row.get("Site", ""),
            "OS":            ir["OS version"]                  if ir is not None else "N/A",
            "Model":         ir["Model"]                       if ir is not None else "N/A",
            "Username":      ir["Username"]                    if ir is not None else "N/A",
            "Last Response": str(ir["Last response (Local time)"]) if ir is not None else "N/A",
        }
        browsers = parse_browsers(str(row["Output"]))
        dual_32, only_32, per_user = detect_issues(browsers)

        issue_sets = [
            ("32-bit Install (Dual — remove 32-bit copy)",    dual_32),
            ("32-bit Install (Only — replace with 64-bit)",   only_32),
            ("Per-User (AppData) Install",                    per_user),
        ]
        for issue_type, installs in issue_sets:
            for b in installs:
                key = (device, b["browser"], b["path"], issue_type)
                if key in seen: continue
                seen.add(key)
                flagged.append({**base,
                    "Browser":       b["browser"],
                    "Issue Type":    issue_type,
                    "Issue Detail":  f"v{b['version']} @ {b['path']}",
                    "Version":       b["version"],
                    "Architecture":  b["arch"],
                    "Install Scope": b["scope"],
                    "Install Path":  b["path"],
                })
    return flagged


BROWSERS_TRACKED = ["Google Chrome", "Microsoft Edge", "Mozilla Firefox"]

def build_browser_sheet_rows(audit_active, inv_active, browser_name):
    """One row per install of browser_name, sorted by version descending."""
    idx = inv_active.set_index("Device name")
    rows = []
    for _, row in audit_active.iterrows():
        device = row["Device"]
        inv = idx.loc[device] if device in idx.index else None
        for b in parse_browsers(str(row["Output"])):
            if b["browser"] != browser_name:
                continue
            is_per_user = b["scope"] == "Per-User" or "appdata" in b["path"].lower()

            # For Firefox, pull the most-recent last-used date across all profiles on this device
            ff_last_used = ""
            if browser_name == "Mozilla Firefox":
                ff_records = parse_firefox_last_used(str(row["Output"]))
                valid_dts = [r["last_used_dt"] for r in ff_records if r["last_used_dt"]]
                if valid_dts:
                    ff_last_used = max(valid_dts).strftime("%d %b %Y  %H:%M")

            entry = {
                "Device":              device,
                "Client":              inv["Customer name"]               if inv is not None and "Customer name"             in inv.index else row.get("Client", ""),
                "Site":                inv["Site name"]                   if inv is not None and "Site name"                 in inv.index else row.get("Site", ""),
                "Version":             b["version"],
                "Architecture":        b["arch"],
                "Install Scope":       b["scope"],
                "Install Path":        b["path"],
                "Is 32-bit":           "Yes" if b["arch"] == "32-bit" else "No",
                "Is Per-User/AppData": "Yes" if is_per_user else "No",
                "Note":                _path_note(b["arch"], b["path"]),
                "OS":                  inv["OS version"]                  if inv is not None else "N/A",
                "Username":            inv["Username"]                    if inv is not None else "N/A",
                "Last Response":       str(inv["Last response (Local time)"]) if inv is not None else "N/A",
            }
            if browser_name == "Mozilla Firefox":
                entry["Last Used"] = ff_last_used
            rows.append(entry)
    # Sort: by device name first so all installs for the same device sit together
    # (e.g. MEDTECH121 Firefox 32-bit and 64-bit are adjacent rows).
    # Within each device: risk rows (32-bit / per-user) before clean, then version desc.
    # Devices that have ANY risk install float to the top of the sheet so they're
    # visible without filtering; fully clean devices follow below.
    device_has_risk = {
        r["Device"]
        for r in rows
        if r["Is 32-bit"] == "Yes" or r["Is Per-User/AppData"] == "Yes"
    }
    def _sort_key(r):
        is_risk_device = r["Device"] in device_has_risk
        is_risk_row    = r["Is 32-bit"] == "Yes" or r["Is Per-User/AppData"] == "Yes"
        ver_parts      = [int(x) for x in r["Version"].split(".") if x.isdigit()]
        # (risk device first, device name, risk row first within device, version desc)
        return (0 if is_risk_device else 1, r["Device"], 0 if is_risk_row else 1, [-p for p in ver_parts])
    rows.sort(key=_sort_key)
    return rows

def build_version_spread(browser_rows):
    """Return list of {Version, Devices, 64-bit Devices, 32-bit Devices, Per-User Devices} sorted by version desc."""
    from collections import defaultdict
    ver_devices  = defaultdict(set)
    ver_64       = defaultdict(set)
    ver_32       = defaultdict(set)
    ver_per_user = defaultdict(set)
    for r in browser_rows:
        v = r["Version"]
        ver_devices[v].add(r["Device"])
        if r["Is 32-bit"] == "No":
            ver_64[v].add(r["Device"])
        if r["Is 32-bit"] == "Yes":
            ver_32[v].add(r["Device"])
        if r["Is Per-User/AppData"] == "Yes":
            ver_per_user[v].add(r["Device"])
    versions = sorted(ver_devices.keys(),
                      key=lambda v: [int(x) for x in v.split(".") if x.isdigit()],
                      reverse=True)
    return [{"Version": v, "Devices": len(ver_devices[v]),
             "64-bit Devices": len(ver_64[v]),
             "32-bit Devices": len(ver_32[v]),
             "Per-User Devices": len(ver_per_user[v])}
            for v in versions]

def browser_overview_counts(all_browser_rows):
    """Summary counts per browser from the combined per-browser row lists."""
    counts = []
    for browser_name, rows in all_browser_rows.items():
        devices     = {r["Device"] for r in rows}
        installs_64 = [r for r in rows if r["Is 32-bit"] == "No"]
        installs_32 = [r for r in rows if r["Is 32-bit"] == "Yes"]
        devices_per = {r["Device"] for r in rows if r["Is Per-User/AppData"] == "Yes"}
        counts.append({
            "Browser":              browser_name,
            "Devices":              len(devices),
            "64-bit Installs":      len(installs_64),
            "32-bit Installs":      len(installs_32),
            "Per-User Installs":    len(devices_per),
        })
    return counts

BROWSER_SHEET_CONFIG = {
    "Google Chrome":   {"emoji": "🟢", "color": COL["green"], "row_alt": COL["row_blue"]},
    "Microsoft Edge":  {"emoji": "🔵", "color": COL["navy"],  "row_alt": COL["row_blue"]},
    "Mozilla Firefox": {"emoji": "🟠", "color": COL["amber"], "row_alt": COL["row_blue"]},
}

def _write_browser_sheet(wb, browser_name, rows):
    cfg = BROWSER_SHEET_CONFIG[browser_name]
    short = browser_name.replace("Google ", "").replace("Mozilla ", "").replace("Microsoft ", "")
    ws = wb.create_sheet(f"{cfg['emoji']} {short}"); ws.sheet_view.showGridLines = False

    is_firefox = browser_name == "Mozilla Firefox"
    hdrs = ["Device", "Version", "Architecture", "Install Scope",
            "Install Path", "Is 32-bit", "Is Per-User/AppData", "Username", "Last Response"]
    if is_firefox:
        hdrs.append("Last Used")

    for c, h in enumerate(hdrs, 1): ws.cell(1, c, h)
    _hdr(ws, 1, len(hdrs), _fill(cfg["color"]))
    for r, rec in enumerate(rows, 2):
        is_risk = rec["Is 32-bit"] == "Yes" or rec["Is Per-User/AppData"] == "Yes"
        alt = _fill(COL["row_risk"]) if is_risk else _fill(COL["row_blue"])
        for c, key in enumerate(hdrs, 1):
            cell = ws.cell(r, c, rec.get(key, "")); cell.font = BODY_FONT
            cell.border = BORDER; cell.fill = alt; cell.alignment = WRAP
        ws.row_dimensions[r].height = 18

    if is_firefox:
        _widths(ws, [24, 18, 14, 16, 64, 10, 18, 24, 20, 22])
    else:
        _widths(ws, [24, 18, 14, 16, 64, 10, 18, 24, 20])
    ws.freeze_panes = "A2"; ws.auto_filter.ref = ws.dimensions
    _write_legend(ws, len(rows) + 3, [
        (COL["row_risk"],  "Risk row — install is 32-bit or per-user/AppData. Review and remediate."),
        (COL["row_blue"],  "Clean row — 64-bit system install, no issues detected."),
    ])


def build_firefox_last_used_rows(audit_active):
    """
    Parse the Firefox Usage Report section from every device's Output field.
    Returns a list of dicts ready to write to the sheet, sorted by last_used_dt desc.
    """
    rows = []
    for _, row in audit_active.iterrows():
        device = row["Device"]
        for rec in parse_firefox_last_used(str(row["Output"])):
            rows.append({
                "Device":       device,
                "Windows User": rec["user"],
                "FF Profile":   rec["profile"],
                "Last Used":    rec["last_used_dt"].strftime("%d %b %Y  %H:%M") if rec["last_used_dt"] else rec["last_used_raw"],
                "_dt":          rec["last_used_dt"] or datetime.min,
            })
    rows.sort(key=lambda r: r["_dt"], reverse=True)
    return rows


def _write_firefox_last_used_sheet(wb, rows):
    """Add a '🦊 Firefox Last Used' sheet to wb."""
    ws = wb.create_sheet("🦊 Firefox Last Used")
    ws.sheet_view.showGridLines = False

    hdrs = ["Device", "Windows User", "FF Profile", "Last Used"]
    for c, h in enumerate(hdrs, 1):
        ws.cell(1, c, h)
    _hdr(ws, 1, len(hdrs), _fill(COL["amber"]))

    now = datetime.now()
    for r, rec in enumerate(rows, 2):
        dt = rec.get("_dt")
        if dt and dt != datetime.min:
            days_ago = (now - dt).days
            row_fill = _fill(COL["row_gold2"]) if days_ago > 90 else _fill(COL["row_blue"])
        else:
            row_fill = _fill(COL["white"])

        for c, key in enumerate(hdrs, 1):
            cell = ws.cell(r, c, rec.get(key, ""))
            cell.font = BODY_FONT; cell.border = BORDER
            cell.fill = row_fill; cell.alignment = WRAP
        ws.row_dimensions[r].height = 18

    _widths(ws, [24, 20, 38, 22])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    _write_legend(ws, len(rows) + 3, [
        (COL["row_blue"],  "Firefox used within the last 90 days."),
        (COL["row_gold2"], "Firefox not used in over 90 days — consider reviewing."),
    ])


def build_report(flagged, not_scanned, stale_inv, inv_active, audit_active, all_browser_rows, output_path, stale_days, client_name=""):
    wb = Workbook()

    # ── Summary ───────────────────────────────────────────────────────────────
    # Single left-aligned column layout — everything reads top to bottom:
    #
    #   Rows  4–9  : Fleet stat tiles           (A=label, B=value)
    #   Row   10   : (blank)
    #   Rows 11–14 : Browser counts table        (A–E: Browser|Devices|64-bit|32-bit|Per-User)
    #   Row   15   : (blank)
    #   Rows 16–19 : Issue breakdown tiles       (A=label, B=value)
    #   Rows 21–22 : (blank / spread title row)
    #   Rows 22+   : Version spread tables       Chrome A–E, Edge G–K, Firefox M–Q
    #
    ws = wb.active; ws.title = "Summary"; ws.sheet_view.showGridLines = False
    ws["A1"] = f"Browser CVE Audit Report — {client_name}" if client_name else "Browser CVE Audit Report"
    ws["A1"].font = Font(name="Arial", bold=True, size=16, color=COL["dark"])
    ws["A2"] = f"Generated {datetime.now().strftime('%d %b %Y')}  |  Devices inactive >{stale_days} days excluded"
    ws["A2"].font = Font(name="Arial", size=10, color="7F8C8D")

    # ── Section 1 — Fleet stat tiles (rows 4–9) ───────────────────────────────
    STAT_ROW = 4
    for i, (label, val, color) in enumerate([
        ("Total Fleet (Inventory)",               len(inv_active) + len(stale_inv), COL["dark"]),
        (f"Stale / Offline (>{stale_days} days)", len(stale_inv),                   COL["gold"]),
        ("Active Devices (Inventory)",            len(inv_active),                   COL["navy"]),
        ("Active Devices Scanned",                len(audit_active),                 COL["green"]),
        ("Devices with Issues",                   len({r["Device"] for r in flagged}), COL["red"]),
        ("Active Devices NOT Scanned",            len(not_scanned),                  COL["amber"]),
    ], STAT_ROW):
        f = _fill(color)
        ws.cell(i, 1, label).font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        ws.cell(i, 1).fill = f; ws.cell(i, 1).border = BORDER
        ws.cell(i, 1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.cell(i, 2, val).font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        ws.cell(i, 2).fill = f; ws.cell(i, 2).border = BORDER; ws.cell(i, 2).alignment = CENTER
        ws.row_dimensions[i].height = 22

    # ── Section 2 — Browser counts table (rows 11–14) ────────────────────────
    # One header row + one coloured row per browser, 5 cols: Browser|Devices|64-bit|32-bit|Per-User
    OV_ROW  = 11
    OV_HDRS = ["Browser", "Devices", "64-bit Installs", "32-bit Installs", "Per-User Installs"]
    for c, h in enumerate(OV_HDRS, 1):
        cell = ws.cell(OV_ROW, c, h)
        cell.fill = _fill(COL["dark"]); cell.font = HDR_FONT
        cell.border = BORDER; cell.alignment = CENTER
    ws.row_dimensions[OV_ROW].height = 22

    browser_counts = browser_overview_counts(all_browser_rows)
    for ri, rec in enumerate(browser_counts, OV_ROW + 1):
        cfg  = BROWSER_SHEET_CONFIG[rec["Browser"]]
        vals = [rec["Browser"], rec["Devices"], rec["64-bit Installs"],
                rec["32-bit Installs"], rec["Per-User Installs"]]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(ri, c, v)
            cell.font  = Font(name="Arial", bold=True, size=10, color="FFFFFF")
            cell.fill  = _fill(cfg["color"]); cell.border = BORDER
            cell.alignment = Alignment(horizontal="left" if c == 1 else "center",
                                       vertical="center", indent=1 if c == 1 else 0)
        ws.row_dimensions[ri].height = 22

    # ── Section 3 — Issue breakdown tiles (rows 16–19) ───────────────────────
    ISSUE_ROW = 16
    all_rows_flat = [r for rows in all_browser_rows.values() for r in rows]
    dual_32_devs  = len({r["Device"] for r in flagged if "Dual" in r["Issue Type"]})
    only_32_devs  = len({r["Device"] for r in flagged if "Only" in r["Issue Type"]})
    per_user_devs = len({r["Device"] for r in flagged if r["Issue Type"] == "Per-User (AppData) Install"})
    for i, (label, val, color) in enumerate([
        ("Devices with dual 32+64-bit installs (remove 32-bit copy)", dual_32_devs,      COL["red"]),
        ("Devices with 32-bit only installs (replace with 64-bit)",   only_32_devs,      COL["amber"]),
        ("Devices with per-user/AppData installs",                    per_user_devs,      COL["navy"]),
        ("Total Chrome/Edge/Firefox installs detected",               len(all_rows_flat), COL["green"]),
    ], ISSUE_ROW):
        f = _fill(color)
        ws.cell(i, 1, label).font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        ws.cell(i, 1).fill = f; ws.cell(i, 1).border = BORDER
        ws.cell(i, 1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.cell(i, 2, val).font = Font(name="Arial", bold=True, size=12, color="FFFFFF")
        ws.cell(i, 2).fill = f; ws.cell(i, 2).border = BORDER; ws.cell(i, 2).alignment = CENTER
        ws.row_dimensions[i].height = 22

    # ── Section 4 — Version spread tables (rows 22+, three browsers side by side) ──
    # 5 cols per table: Version | Devices | 64-bit | 32-bit | Per-User
    # 1 blank-col gap between tables.
    # Spread starts at col A — no collision with other sections (which only use rows 4–19).
    SPREAD_NCOLS     = 5
    SPREAD_GAP       = 1
    SPREAD_HDRS      = ["Version", "Devices", "64-bit", "32-bit", "Per-User"]
    SPREAD_ROW_START = ISSUE_ROW + 4 + 1   # rows 16–19 issues + 1 blank + 1 title = row 22

    for bi, (browser_name, rows) in enumerate(all_browser_rows.items()):
        cfg   = BROWSER_SHEET_CONFIG[browser_name]
        short = browser_name.replace("Google ", "").replace("Mozilla ", "").replace("Microsoft ", "")
        col0  = 1 + bi * (SPREAD_NCOLS + SPREAD_GAP)   # 1, 7, 13

        # Browser title row — merged across all 5 cols
        title_row = SPREAD_ROW_START - 1
        ws.merge_cells(start_row=title_row, start_column=col0,
                       end_row=title_row,   end_column=col0 + SPREAD_NCOLS - 1)
        tc = ws.cell(title_row, col0, f"{cfg['emoji']} {short} — Version Spread")
        tc.font      = Font(name="Arial", bold=True, size=10, color="FFFFFF")
        tc.fill      = _fill(cfg["color"]); tc.border = BORDER
        tc.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[title_row].height = 20

        # Column headers
        for ci, h in enumerate(SPREAD_HDRS, col0):
            cell = ws.cell(SPREAD_ROW_START, ci, h)
            cell.fill = _fill(COL["dark"]); cell.font = HDR_FONT
            cell.border = BORDER; cell.alignment = CENTER
        ws.row_dimensions[SPREAD_ROW_START].height = 20

        # Data rows
        spread = build_version_spread(rows)
        for ri, sv in enumerate(spread, SPREAD_ROW_START + 1):
            if sv["32-bit Devices"] > 0:
                row_fill = _fill(COL["row_only32"])
            elif sv["Per-User Devices"] > 0:
                row_fill = _fill(COL["row_peruser"])
            else:
                row_fill = _fill(COL["row_blue"])
            vals = [sv["Version"], sv["Devices"], sv["64-bit Devices"],
                    sv["32-bit Devices"], sv["Per-User Devices"]]
            for ci, v in enumerate(vals, col0):
                cell = ws.cell(ri, ci, v); cell.font = BODY_FONT
                cell.border = BORDER; cell.fill = row_fill; cell.alignment = CENTER
            ws.row_dimensions[ri].height = 18

        # Column widths
        for ci, w in zip(range(col0, col0 + SPREAD_NCOLS), [18, 10, 10, 10, 11]):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.column_dimensions[get_column_letter(col0 + SPREAD_NCOLS)].width = 2

    # Global column widths for cols A–B and overview cols C–E
    ws.column_dimensions["A"].width = 54
    ws.column_dimensions["B"].width = 10
    for ci, w in enumerate([20, 16, 16, 18], 3):   # C=Browser, D=Devices, E=64-bit, F=32-bit, G=Per-User
        ws.column_dimensions[get_column_letter(ci)].width = w

    # ── ⚠ Devices with Issues ─────────────────────────────────────────────────
    ws2 = wb.create_sheet("⚠ Devices with Issues"); ws2.sheet_view.showGridLines = False
    hdrs = ["Device", "Client", "Site", "Browser", "Issue Type", "Version",
            "Architecture", "Install Scope", "Install Path", "OS", "Model", "Username", "Last Response"]
    # Colour by issue type so it's visually scannable without filtering
    _ISSUE_FILL = {
        "32-bit Install (Dual — remove 32-bit copy)":  COL["row_dual32"],
        "32-bit Install (Only — replace with 64-bit)": COL["row_only32"],
        "Per-User (AppData) Install":                  COL["row_peruser"],
    }
    for c, h in enumerate(hdrs, 1): ws2.cell(1, c, h)
    _hdr(ws2, 1, len(hdrs), _fill(COL["dark"]))
    sorted_flagged = sorted(flagged, key=lambda x: (x["Client"], x["Site"], x["Device"], x["Browser"], x["Issue Type"]))
    for r, rec in enumerate(sorted_flagged, 2):
        row_fill = _fill(_ISSUE_FILL.get(rec.get("Issue Type", ""), COL["row_risk"]))
        for c, key in enumerate(hdrs, 1):
            cell = ws2.cell(r, c, rec.get(key, "")); cell.font = BODY_FONT
            cell.border = BORDER; cell.fill = row_fill; cell.alignment = WRAP
        ws2.row_dimensions[r].height = 18
    _widths(ws2, [22, 14, 22, 20, 38, 16, 14, 14, 58, 36, 26, 26, 20]); ws2.freeze_panes = "A2"
    ws2.auto_filter.ref = ws2.dimensions
    _write_legend(ws2, len(sorted_flagged) + 3, [
        (COL["row_dual32"],  "32-bit Install (Dual) — same browser has both 32-bit and 64-bit installed. Remove the 32-bit copy."),
        (COL["row_only32"],  "32-bit Install (Only) — browser only present as 32-bit, no 64-bit version. Replace with 64-bit installer."),
        (COL["row_peruser"], "Per-User (AppData) Install — browser installed under user profile, not system-wide. Reinstall to Program Files."),
    ])

    # ── Per-browser sheets ────────────────────────────────────────────────────
    for browser_name in BROWSERS_TRACKED:
        _write_browser_sheet(wb, browser_name, all_browser_rows[browser_name])

    # ── Firefox Last Used sheet ───────────────────────────────────────────────
    ff_last_used_rows = build_firefox_last_used_rows(audit_active)
    if ff_last_used_rows:
        _write_firefox_last_used_sheet(wb, ff_last_used_rows)

    ws3 = wb.create_sheet("🔍 Not Scanned (Active)"); ws3.sheet_view.showGridLines = False
    ns_h = ["Device Name", "Customer", "Site", "Device Type", "OS Version", "Username", "Last Response", "Manufacturer", "Model"]
    for c, h in enumerate(ns_h, 1): ws3.cell(1, c, h)
    _hdr(ws3, 1, len(ns_h), _fill(COL["amber"]))
    for r, (_, rec) in enumerate(not_scanned.iterrows(), 2):
        alt = _fill(COL["row_blue"])
        vals = [rec["Device name"], rec["Customer name"], rec["Site name"], rec["Device type"],
                rec["OS version"], rec["Username"], str(rec["Last response (Local time)"]), rec["Manufacturer"], rec["Model"]]
        for c, v in enumerate(vals, 1):
            cell = ws3.cell(r, c, v); cell.font = BODY_FONT
            cell.border = BORDER; cell.fill = alt; cell.alignment = WRAP
        ws3.row_dimensions[r].height = 18
    _widths(ws3, [22, 14, 24, 14, 38, 26, 20, 16, 34]); ws3.freeze_panes = "A2"

    ws4 = wb.create_sheet(f"🕐 Stale >{stale_days}d (Excluded)"); ws4.sheet_view.showGridLines = False
    st_h = ["Device Name", "Customer", "Site", "Device Type", "OS Version", "Username", "Last Response", "Days Offline", "Manufacturer", "Model"]
    for c, h in enumerate(st_h, 1): ws4.cell(1, c, h)
    _hdr(ws4, 1, len(st_h), _fill(COL["gold"]))
    for r, (_, rec) in enumerate(stale_inv.sort_values("days_since", ascending=False).iterrows(), 2):
        alt = _fill(COL["row_gold"])
        days = int(rec["days_since"]) if pd.notna(rec["days_since"]) else "?"
        vals = [rec["Device name"], rec["Customer name"], rec["Site name"], rec["Device type"],
                rec["OS version"], rec["Username"], str(rec["Last response (Local time)"]), days, rec["Manufacturer"], rec["Model"]]
        for c, v in enumerate(vals, 1):
            cell = ws4.cell(r, c, v); cell.font = BODY_FONT
            cell.border = BORDER; cell.fill = alt; cell.alignment = WRAP
        ws4.row_dimensions[r].height = 18
    _widths(ws4, [22, 14, 24, 14, 38, 26, 20, 18, 16, 34]); ws4.freeze_panes = "A2"

    wb.save(output_path)
    return len({r["Device"] for r in flagged}), len(not_scanned), len(stale_inv)



def get_output_path():
    """Return a writable report path. Handles redirected/missing Desktop folders."""
    filename = f"Browser_CVE_Report_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.xlsx"
    candidates = [
        Path.home() / "Desktop",
        Path.home() / "OneDrive" / "Desktop",
        Path.home() / "Downloads",
        Path.cwd(),
    ]
    for folder in candidates:
        try:
            folder.mkdir(parents=True, exist_ok=True)
            test_file = folder / ".browser_cve_write_test"
            test_file.write_text("test", encoding="utf-8")
            test_file.unlink(missing_ok=True)
            return folder / filename
        except Exception:
            continue
    return Path.cwd() / filename


# ═══════════════════════════════════════════════════════════════════════════════
#  GUI — three big buttons, a log, and a run button
# ═══════════════════════════════════════════════════════════════════════════════
class App(tk.Tk):
    BG      = "#1E2530"
    CARD    = "#252D3A"
    ACCENT  = "#3B82F6"
    ACCENTH = "#2563EB"
    TEXT    = "#E2E8F0"
    SUBTEXT = "#94A3B8"
    SUCCESS = "#22C55E"
    ERROR   = "#EF4444"
    WARN    = "#F59E0B"
    BORDERC = "#374151"

    def __init__(self):
        super().__init__()
        self.title("Browser CVE Audit v6")
        self.configure(bg=self.BG)
        self.resizable(False, False)

        self.task_files = []
        self.inv_path   = ""
        self.out_path   = ""
        self.stale_days = tk.IntVar(value=30)

        self._build()
        w, h = 540, 550
        self.update_idletasks()
        x = (self.winfo_screenwidth()  - w) // 2
        y = (self.winfo_screenheight() - h) // 2
        self.geometry(f"{w}x{h}+{x}+{y}")

    # ── helpers ───────────────────────────────────────────────────────────────
    def _lbl(self, p, text, size=10, bold=False, color=None):
        return tk.Label(p, text=text,
                        font=("Segoe UI", size, "bold" if bold else "normal"),
                        fg=color or self.TEXT, bg=p.cget("bg"))

    def _big_btn(self, parent, text, sub, cmd, done_var):
        """Card-style button with a status sub-label."""
        card = tk.Frame(parent, bg=self.CARD, cursor="hand2",
                        highlightthickness=1, highlightbackground=self.BORDERC)
        card.pack(fill="x", padx=20, pady=6)
        inner = tk.Frame(card, bg=self.CARD)
        inner.pack(fill="x", padx=14, pady=10)
        tk.Label(inner, text=text, font=("Segoe UI", 10, "bold"),
                 fg=self.TEXT, bg=self.CARD, anchor="w").pack(fill="x")
        status = tk.Label(inner, textvariable=done_var, font=("Segoe UI", 8),
                          fg=self.SUBTEXT, bg=self.CARD, anchor="w")
        status.pack(fill="x")
        for widget in (card, inner, status):
            widget.bind("<Button-1>", lambda e: cmd())
            widget.bind("<Enter>",    lambda e: card.config(highlightbackground=self.ACCENT))
            widget.bind("<Leave>",    lambda e: card.config(highlightbackground=self.BORDERC))
        return card

    # ── layout ────────────────────────────────────────────────────────────────
    def _build(self):
        # Header
        hdr = tk.Frame(self, bg="#141920", pady=16)
        hdr.pack(fill="x")
        self._lbl(hdr, "🔍  Browser CVE Audit", size=15, bold=True).pack()
        self._lbl(hdr, "Select input files, choose where to save, then generate",
                  size=9, color=self.SUBTEXT).pack(pady=(2, 0))

        # Step labels
        steps = tk.Frame(self, bg=self.BG)
        steps.pack(fill="x", padx=20, pady=(14, 2))
        self._lbl(steps, "STEP 1 & 2 — Select input files", size=8,
                  color=self.SUBTEXT).pack(anchor="w")

        # CSV button
        self.csv_label = tk.StringVar(value="No files selected")
        self._big_btn(self, "📄  Task Report CSV(s)", "",
                      self._pick_csvs, self.csv_label)

        # Inventory button
        self.inv_label = tk.StringVar(value="No file selected")
        self._big_btn(self, "📋  Device Inventory XLSX", "",
                      self._pick_inv, self.inv_label)

        # Output file button
        self.out_label = tk.StringVar(value="No save location selected")
        self._big_btn(self, "💾  Save Report As", "",
                      self._pick_output, self.out_label)

        # Stale days row
        stale_row = tk.Frame(self, bg=self.BG)
        stale_row.pack(fill="x", padx=20, pady=(4, 0))
        self._lbl(stale_row, "Exclude devices offline for more than", size=9).pack(side="left")
        tk.Spinbox(stale_row, from_=1, to=365, textvariable=self.stale_days,
                   width=4, font=("Segoe UI", 9), bg="#2D3748", fg=self.TEXT,
                   buttonbackground="#374151", relief="flat",
                   highlightthickness=1, highlightbackground=self.BORDERC).pack(
            side="left", padx=6)
        self._lbl(stale_row, "days", size=9).pack(side="left")

        # Log
        log_frame = tk.Frame(self, bg=self.BG)
        log_frame.pack(fill="x", padx=20, pady=(12, 0))
        self._lbl(log_frame, "STEP 3 — Run", size=8, color=self.SUBTEXT).pack(anchor="w")
        self.log = tk.Text(log_frame, height=5, font=("Consolas", 9),
                           bg="#0F1419", fg="#A8B2C1", relief="flat",
                           state="disabled", wrap="word",
                           highlightthickness=1, highlightbackground=self.BORDERC)
        self.log.pack(fill="x", pady=(4, 0))
        self.log.tag_config("ok",   foreground=self.SUCCESS)
        self.log.tag_config("err",  foreground=self.ERROR)
        self.log.tag_config("warn", foreground=self.WARN)
        self.log.tag_config("dim",  foreground=self.SUBTEXT)

        # Generate button
        foot = tk.Frame(self, bg=self.BG, pady=14)
        foot.pack(fill="x")
        self.run_btn = tk.Button(
            foot, text="▶  Generate Report", command=self._run,
            font=("Segoe UI", 11, "bold"), bg=self.ACCENT, fg=self.TEXT,
            activebackground=self.ACCENTH, activeforeground=self.TEXT,
            relief="flat", cursor="hand2", padx=20, pady=8)
        self.run_btn.pack()
        self.run_btn.bind("<Enter>", lambda e: self.run_btn.config(bg=self.ACCENTH))
        self.run_btn.bind("<Leave>", lambda e: self.run_btn.config(bg=self.ACCENT))
        self.progress = ttk.Progressbar(foot, mode="indeterminate", length=260)
        self.progress.pack(pady=(8, 0))

    # ── file pickers ──────────────────────────────────────────────────────────
    def _pick_csvs(self):
        files = filedialog.askopenfilenames(
            title="Select Task Report CSV(s)",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")])
        if files:
            self.task_files = list(files)
            names = ", ".join(Path(f).name for f in files)
            self.csv_label.set(f"✔  {len(files)} file(s): {names}")

    def _pick_inv(self):
        f = filedialog.askopenfilename(
            title="Select Device Inventory XLSX",
            filetypes=[("Excel files", "*.xlsx *.xls"), ("All", "*.*")])
        if f:
            self.inv_path = f
            self.inv_label.set(f"✔  {Path(f).name}")

    def _pick_output(self):
        default_name = f"Browser_CVE_Report_{datetime.now().strftime('%Y-%m-%d_%H%M%S')}.xlsx"
        f = filedialog.asksaveasfilename(
            title="Choose where to save the Browser CVE report",
            defaultextension=".xlsx",
            initialfile=default_name,
            filetypes=[("Excel workbook", "*.xlsx"), ("All files", "*.*")],
        )
        if f:
            if not f.lower().endswith(".xlsx"):
                f += ".xlsx"
            self.out_path = f
            self.out_label.set(f"✔  {f}")

    # ── logging ───────────────────────────────────────────────────────────────
    def _log(self, msg, tag=""):
        self.log.config(state="normal")
        self.log.insert("end", msg + "\n", tag)
        self.log.see("end")
        self.log.config(state="disabled")

    # ── run ───────────────────────────────────────────────────────────────────
    def _run(self):
        if not self.task_files:
            messagebox.showerror("Missing", "Please select at least one Task Report CSV.")
            return
        if not self.inv_path:
            messagebox.showerror("Missing", "Please select the Device Inventory XLSX.")
            return
        if not self.out_path:
            self._pick_output()
            if not self.out_path:
                messagebox.showerror("Missing", "Please choose where to save the report.")
                return

        self.log.config(state="normal"); self.log.delete("1.0", "end"); self.log.config(state="disabled")
        self.run_btn.config(state="disabled")
        self.progress.start(12)
        threading.Thread(target=self._worker, daemon=True).start()

    def _worker(self):
        try:
            stale_days = self.stale_days.get()
            out_path = self.out_path

            self._log(f"Task CSV(s): {len(self.task_files)} selected", "dim")
            for file in self.task_files:
                self._log(f"  CSV: {file}", "dim")
            self._log(f"Inventory: {self.inv_path}", "dim")
            self._log(f"Save as: {out_path}", "dim")

            self._log("Loading inventory…", "dim")
            inv_active, stale_inv, client_name = load_inventory(self.inv_path, stale_days)
            self._log(f"  Active: {len(inv_active)}  |  Stale: {len(stale_inv)}", "dim")

            self._log(f"Loading {len(self.task_files)} CSV(s)…", "dim")
            audit = load_audit(self.task_files)
            self._log(f"  Audit records: {len(audit)}", "dim")

            # Inventory is the source of truth for the selected client/scope.
            # Only devices present in the active inventory are included in scanned/issue/browser counts.
            active_names = set(inv_active["Device name"])
            audit_device_names = set(audit["Device"])
            audit_active = audit[audit["Device"].isin(active_names)].copy()
            scanned_names = set(audit_active["Device"])
            not_scanned = inv_active[
                inv_active["Device name"].isin(active_names - scanned_names)
            ].sort_values(["Site name", "Device name"])
            ignored_not_inventory = sorted(audit_device_names - active_names)
            if ignored_not_inventory:
                self._log(f"  Ignored scanned CSV devices not in inventory: {len(ignored_not_inventory)}", "warn")
            self._log(f"  In-scope scanned devices: {len(audit_active)}", "dim")

            self._log("Checking browser installs…", "dim")
            flagged = build_flagged(audit_active, inv_active)
            all_browser_rows = {b: build_browser_sheet_rows(audit_active, inv_active, b) for b in BROWSERS_TRACKED}

            self._log("Writing report…", "dim")
            Path(out_path).parent.mkdir(parents=True, exist_ok=True)
            n_f, n_ns, n_st = build_report(
                flagged, not_scanned, stale_inv, inv_active, audit_active,
                all_browser_rows,
                out_path, stale_days, client_name)

            all_rows_flat = [r for rows in all_browser_rows.values() for r in rows]
            ff_last_used_count = sum(
                len(parse_firefox_last_used(str(row["Output"])))
                for _, row in audit_active.iterrows()
            )
            self._log(f"\n✔  Saved to: {out_path}", "ok")
            self._log(f"   Issues found    : {n_f} devices / {len(flagged)} issue rows",  "warn" if n_f  else "ok")
            self._log(f"   Browser installs: {len(all_rows_flat)}", "dim")
            n_dual   = len({r["Device"] for r in flagged if "Dual" in r["Issue Type"]})
            n_only32 = len({r["Device"] for r in flagged if "Only" in r["Issue Type"]})
            n_per    = len({r["Device"] for r in flagged if r["Issue Type"] == "Per-User (AppData) Install"})
            self._log(f"   32-bit dual     : {n_dual} devices", "warn" if n_dual else "ok")
            self._log(f"   32-bit only     : {n_only32} devices", "warn" if n_only32 else "ok")
            self._log(f"   Per-user        : {n_per} devices", "warn" if n_per else "ok")
            self._log(f"   Not scanned     : {n_ns}", "warn" if n_ns else "ok")
            self._log(f"   Stale excluded  : {n_st}", "dim")
            self._log(f"   FF last-used    : {ff_last_used_count} profile record(s)", "dim")

            self.after(0, lambda: self._offer_open(out_path))

        except Exception as exc:
            import traceback
            self._log(f"\n✘  {exc}", "err")
            self._log(traceback.format_exc(), "err")
        finally:
            self.after(0, self._done)

    def _done(self):
        self.progress.stop()
        self.run_btn.config(state="normal")

    def _offer_open(self, path):
        if messagebox.askyesno("Done!", f"Report saved to:\n{path}\n\nOpen it now?"):
            import os, subprocess, sys
            if os.name == "nt":
                os.startfile(path)
            elif sys.platform == "darwin":
                subprocess.Popen(["open", path])
            else:
                subprocess.Popen(["xdg-open", path])


if __name__ == "__main__":
    App().mainloop()